import uuid
import csv
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from pydantic import BaseModel

from app.dependencies import get_db, require_role
from app.models.user import User
from app.models.job import Job, Application
from app.services.email_service import send_listing_submitted_email, send_bulk_upload_email

_VALID_JOB_TYPES = {"full_time", "part_time", "contract"}
_VALID_EXP_LEVELS = {"entry", "mid", "senior", "lead"}
_VALID_REMOTE_TYPES = {"onsite", "remote", "hybrid"}
_BULK_MIN_ROWS = 25
_TEMPLATE_HEADERS = [
    "title", "company", "location", "description",
    "job_type", "experience_level", "remote_type",
    "salary_min", "salary_max", "skills_required", "requirements", "vacancies",
]

router = APIRouter(prefix="/provider", tags=["provider"])


class JobCreateRequest(BaseModel):
    title: str
    company: str
    location: str
    description: str
    requirements: list[str] = []
    skills_required: list[str] = []
    salary_min: int | None = None
    salary_max: int | None = None
    job_type: str = "full_time"
    experience_level: str = "mid"
    remote_type: str = "onsite"
    vacancies: int | None = 1


class JobUpdateRequest(BaseModel):
    title: str | None = None
    company: str | None = None
    location: str | None = None
    description: str | None = None
    requirements: list[str] | None = None
    skills_required: list[str] | None = None
    salary_min: int | None = None
    salary_max: int | None = None
    job_type: str | None = None
    experience_level: str | None = None
    remote_type: str | None = None
    vacancies: int | None = None


def _job_to_dict(job: Job, applicant_count: int = 0) -> dict:
    return {
        "id": str(job.id),
        "title": job.title,
        "company": job.company,
        "location": job.location,
        "description": job.description,
        "requirements": job.requirements,
        "skills_required": job.skills_required,
        "salary_min": job.salary_min,
        "salary_max": job.salary_max,
        "currency": job.currency,
        "job_type": job.job_type,
        "experience_level": job.experience_level,
        "remote_type": job.remote_type,
        "vacancies": job.vacancies,
        "review_status": job.review_status,
        "applicants_access": job.applicants_access,
        "is_active": job.is_active,
        "posted_at": job.posted_at.isoformat() if job.posted_at else None,
        "applicant_count": applicant_count,
    }


@router.post("/jobs", status_code=status.HTTP_201_CREATED)
async def create_job(
    data: JobCreateRequest,
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    job = Job(
        title=data.title,
        company=data.company,
        location=data.location,
        description=data.description,
        requirements=data.requirements,
        skills_required=data.skills_required,
        salary_min=data.salary_min,
        salary_max=data.salary_max,
        job_type=data.job_type,
        experience_level=data.experience_level,
        remote_type=data.remote_type,
        vacancies=data.vacancies,
        source="provider",
        is_active=False,
        review_status="pending",
        posted_by=current_user.id,
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)
    await send_listing_submitted_email(current_user.email, current_user.name or "", job.title, job.company or "")
    return _job_to_dict(job)


@router.get("/jobs")
async def list_my_jobs(
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Job).where(Job.posted_by == current_user.id).order_by(Job.posted_at.desc())
    )
    jobs = result.scalars().all()

    counts_result = await db.execute(
        select(Application.job_id, func.count(Application.id))
        .where(Application.job_id.in_([j.id for j in jobs]))
        .group_by(Application.job_id)
    )
    counts = {str(row[0]): row[1] for row in counts_result.all()}

    return [_job_to_dict(j, counts.get(str(j.id), 0)) for j in jobs]


@router.get("/jobs/bulk-template")
async def get_bulk_template():
    """Return an XLSX template file for bulk job upload."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"

        header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(_TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        samples = [
            ["Senior React Developer", "Acme Corp", "Bangalore",
             "We are seeking an experienced React developer to join our growing team.",
             "full_time", "senior", "hybrid", 800000, 1200000,
             "React,TypeScript,Redux", "3+ years React experience|Strong communication skills", 2],
            ["Python Backend Engineer", "Tech Startup", "Mumbai",
             "Build scalable REST APIs for our fintech platform.",
             "full_time", "mid", "remote", 600000, 900000,
             "Python,FastAPI,PostgreSQL", "2+ years Python experience|API design knowledge", 1],
            ["UI/UX Designer", "Design Agency", "Delhi",
             "Create intuitive user experiences for our clients.",
             "contract", "mid", "onsite", 500000, 700000,
             "Figma,Adobe XD,Sketch", "Portfolio required|3+ years UX design experience", 3],
        ]
        for r, row in enumerate(samples, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        col_widths = [28, 20, 20, 50, 14, 17, 12, 13, 13, 32, 42, 10]
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

        ws2 = wb.create_sheet("Valid Values")
        notes = [
            ["Field", "Valid Values", "Notes"],
            ["job_type", "full_time | part_time | contract", ""],
            ["experience_level", "entry | mid | senior | lead", ""],
            ["remote_type", "onsite | remote | hybrid", ""],
            ["skills_required", "comma-separated", "e.g. React,TypeScript"],
            ["requirements", "pipe-separated (|)", "e.g. 3 yrs exp|Strong communication"],
            ["salary_min / salary_max", "integer (annual INR)", "e.g. 600000 for ₹6 LPA"],
            ["vacancies", "integer ≥ 1", "Number of open positions"],
        ]
        for row in notes:
            ws2.append(row)

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="proaicv_jobs_template.xlsx"'},
        )
    except ImportError:
        # openpyxl not installed — fall back to CSV
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(_TEMPLATE_HEADERS)
        w.writerow(["Senior React Developer", "Acme Corp", "Bangalore",
                    "We are seeking an experienced React developer.", "full_time", "senior",
                    "hybrid", 800000, 1200000, "React,TypeScript", "3+ years React experience", 2])
        return Response(
            content=out.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="proaicv_jobs_template.csv"'},
        )


@router.post("/jobs/bulk")
async def bulk_upload_jobs(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    """Bulk upload jobs from a CSV or XLSX file. Minimum 25 rows required."""
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported.")

    content = await file.read()

    # Parse rows
    rows: list[dict] = []
    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(status_code=400, detail="The uploaded file is empty.")
            headers = [str(h).strip().lower() if h else "" for h in all_rows[0]]
            for row in all_rows[1:]:
                if all(cell is None for cell in row):
                    continue
                rows.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))})
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    if len(rows) < _BULK_MIN_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"Bulk upload requires at least {_BULK_MIN_ROWS} job entries. "
                   f"Your file has {len(rows)} data rows. For fewer jobs, use the single job form.",
        )

    success_count = 0
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):
        try:
            title = row.get("title", "").strip()
            company = row.get("company", "").strip()
            location = row.get("location", "").strip()
            description = row.get("description", "").strip()
            row_errors = []
            if not title: row_errors.append("title is required")
            if not company: row_errors.append("company is required")
            if not location: row_errors.append("location is required")
            if not description: row_errors.append("description is required")

            job_type = (row.get("job_type", "") or "full_time").strip().lower()
            if job_type not in _VALID_JOB_TYPES:
                row_errors.append(f"job_type '{job_type}' invalid — use: {', '.join(_VALID_JOB_TYPES)}")
                job_type = "full_time"

            exp_level = (row.get("experience_level", "") or "mid").strip().lower()
            if exp_level not in _VALID_EXP_LEVELS:
                row_errors.append(f"experience_level '{exp_level}' invalid — use: {', '.join(_VALID_EXP_LEVELS)}")
                exp_level = "mid"

            remote_type = (row.get("remote_type", "") or "onsite").strip().lower()
            if remote_type not in _VALID_REMOTE_TYPES:
                row_errors.append(f"remote_type '{remote_type}' invalid — use: {', '.join(_VALID_REMOTE_TYPES)}")
                remote_type = "onsite"

            salary_min = None
            salary_max = None
            try:
                v = row.get("salary_min", "").strip()
                if v:
                    salary_min = int(float(v))
            except (ValueError, TypeError):
                row_errors.append("salary_min must be a number")

            try:
                v = row.get("salary_max", "").strip()
                if v:
                    salary_max = int(float(v))
            except (ValueError, TypeError):
                row_errors.append("salary_max must be a number")

            vacancies = 1
            try:
                v = row.get("vacancies", "").strip()
                if v:
                    vacancies = int(v)
                    if vacancies < 1:
                        row_errors.append("vacancies must be ≥ 1")
                        vacancies = 1
            except (ValueError, TypeError):
                row_errors.append("vacancies must be a whole number")

            if row_errors:
                errors.append({"row": i, "message": "; ".join(row_errors)})
                continue

            skills = [s.strip() for s in row.get("skills_required", "").split(",") if s.strip()]
            reqs = [r.strip() for r in row.get("requirements", "").split("|") if r.strip()]

            job = Job(
                title=title, company=company, location=location, description=description,
                requirements=reqs, skills_required=skills,
                salary_min=salary_min, salary_max=salary_max,
                job_type=job_type, experience_level=exp_level, remote_type=remote_type,
                vacancies=vacancies, source="provider",
                is_active=False, review_status="pending",
                posted_by=current_user.id,
            )
            db.add(job)
            success_count += 1
        except Exception as exc:
            errors.append({"row": i, "message": str(exc)})

    if success_count > 0:
        await db.commit()

    await send_bulk_upload_email(
        current_user.email, current_user.name or "",
        submitted=len(rows), success_count=success_count, errors=errors,
    )

    return {
        "submitted": len(rows),
        "success_count": success_count,
        "pending_review": success_count,
        "error_count": len(errors),
        "errors": errors,
    }


@router.put("/jobs/{job_id}")
async def update_job(
    job_id: uuid.UUID,
    data: JobUpdateRequest,
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Job).where(Job.id == job_id))
    job = result.scalar_one_or_none()
    if not job or job.posted_by != current_user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")
    if job.review_status == "approved":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Approved listings cannot be edited. Contact support.",
        )

    for field, value in data.model_dump(exclude_none=True).items():
        setattr(job, field, value)
    # Reset to pending review after edit
    job.review_status = "pending"
    job.is_active = False
    await db.commit()
    await db.refresh(job)
    return _job_to_dict(job)


@router.delete("/jobs/{job_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_job(
    job_id: uuid.UUID,
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Job).where(Job.id == job_id))
    job = result.scalar_one_or_none()
    if not job or job.posted_by != current_user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")
    await db.delete(job)
    await db.commit()


@router.post("/jobs/{job_id}/request-applicant-access", status_code=200)
async def request_applicant_access(
    job_id: uuid.UUID,
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    """Provider requests admin approval to view applicant details for a listing."""
    result = await db.execute(select(Job).where(Job.id == job_id))
    job = result.scalar_one_or_none()
    if not job or job.posted_by != current_user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")
    if job.review_status != "approved":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Listing must be approved before requesting applicant access.")
    if job.applicants_access in ("pending", "approved"):
        return {"applicants_access": job.applicants_access, "message": "Request already submitted."}
    job.applicants_access = "pending"
    await db.commit()
    return {"applicants_access": "pending", "message": "Access request submitted. Admin will review shortly."}


@router.get("/jobs/{job_id}/applicants")
async def get_applicants(
    job_id: uuid.UUID,
    current_user: User = Depends(require_role("job_provider")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Job).where(Job.id == job_id))
    job = result.scalar_one_or_none()
    if not job or job.posted_by != current_user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job not found")

    if job.applicants_access != "approved":
        access = job.applicants_access or "not_requested"
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={"code": "applicants_access_required", "access": access},
        )

    apps_result = await db.execute(
        select(Application, User)
        .join(User, Application.user_id == User.id)
        .where(Application.job_id == job_id)
        .order_by(Application.applied_at.desc())
    )
    rows = apps_result.all()
    return [
        {
            "application_id": str(app.id),
            "applicant_name": user.name,
            "applicant_email": user.email,
            "applied_at": app.applied_at.isoformat(),
            "status": app.status,
        }
        for app, user in rows
    ]
